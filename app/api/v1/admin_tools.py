import asyncio
import io
import logging
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import desc, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import lazyload, selectinload

from core.config import (
    ALLOWED_BOOK_EXT,
    ALLOWED_IMAGE_EXT,
    BOOK_CATEGORIES,
    MAX_BOOK_SIZE,
    MAX_IMAGE_SIZE,
    MEDIA_ROOT,
    QUESTION_IMAGES_DIR_NAME,
    settings,
)
from db.session import get_db
from models.attempt import QuizAttempt
from models.book import Book
from models.contest import Contest, ContestAttempt, ContestQuestion
from models.module import Module
from models.question import CorrectOption, Question
from models.quiz import Quiz
from models.shop import BookOrder, OrderStatus, ShopBook, ShopSettings
from models.telegram_user import TelegramUser
from models.topic import Topic
from services.notifications import broadcast, notify_new_contest, notify_new_quiz

router = APIRouter(prefix="/admin-tools", tags=["admin-tools"])
templates = Jinja2Templates(directory="templates")

BOOKS_DIR = MEDIA_ROOT / "books"


@router.get("/leaderboard", response_class=HTMLResponse)
async def admin_leaderboard_page(request: Request):
    return templates.TemplateResponse("admin_leaderboard.html", {"request": request})


_PERIOD_LABELS = {"day": "Bugun", "week": "Hafta", "month": "Oy", "all": "Butun_davr"}


@router.get("/leaderboard/export")
async def leaderboard_export(period: str = "all", db: AsyncSession = Depends(get_db)):
    """Export the global rating (for the selected period) as an .xlsx file.

    Mirrors the aggregation used by the WebApp leaderboard endpoint
    (api.v1.webapp.leaderboard): points = SUM(score) * 2, tiebreak by total
    time ascending.
    """
    from api.v1.webapp import _period_since

    total_points = (func.sum(QuizAttempt.score) * 2).label("points")
    total_time = func.sum(QuizAttempt.time_taken_seconds).label("time_taken")
    attempts_n = func.count(QuizAttempt.id).label("attempts_n")

    stmt = (
        select(
            TelegramUser.first_name,
            TelegramUser.last_name,
            TelegramUser.username,
            TelegramUser.telegram_id,
            TelegramUser.phone,
            total_points,
            total_time,
            attempts_n,
        )
        .join(QuizAttempt, QuizAttempt.user_id == TelegramUser.id)
        .group_by(TelegramUser.id)
        .order_by(desc("points"), "time_taken")
    )

    since = _period_since(period)
    if since is not None:
        stmt = stmt.where(QuizAttempt.createdAt >= since)

    rows = (await db.execute(stmt)).all()

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Reyting"

    headers = [
        "O'rin", "Ism Familya", "Username", "Telegram ID", "Telefon",
        "Ballar", "Urinishlar", "Umumiy vaqt (mm:ss)",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="6C63F6", end_color="6C63F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows, start=1):
        full = " ".join(filter(None, [r.first_name, r.last_name])) or ""
        secs = int(r.time_taken or 0)
        mmss = f"{secs // 60:02d}:{secs % 60:02d}"
        ws.append([
            i,
            full,
            r.username or "",
            r.telegram_id,
            r.phone or "",
            int(r.points or 0),
            int(r.attempts_n or 0),
            mmss,
        ])

    widths = [6, 28, 20, 16, 18, 12, 12, 18]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    label = _PERIOD_LABELS.get(period, "reyting")
    filename = f"reyting_{label}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/builder", response_class=HTMLResponse)
async def builder_page(request: Request, db: AsyncSession = Depends(get_db)):
    modules = (await db.execute(select(Module).order_by(Module.order, Module.id))).scalars().all()
    topics = (await db.execute(select(Topic).order_by(Topic.order, Topic.id))).scalars().all()
    return templates.TemplateResponse(
        "admin_builder.html",
        {
            "request": request,
            "modules": [{"id": m.id, "title": m.title} for m in modules],
            "topics": [{"id": t.id, "title": t.title, "module_id": t.module_id} for t in topics],
        },
    )


@router.post("/builder/save")
async def builder_save(payload: dict[str, Any], db: AsyncSession = Depends(get_db)):
    # ── Module (optional but recommended) ──────────────────────────
    module_id = payload.get("module_id")
    new_module_title = (payload.get("new_module_title") or "").strip()

    if new_module_title:
        module = Module(title=new_module_title, is_active=True, order=0)
        db.add(module)
        await db.flush()
        module_id = module.id
    elif module_id:
        module_id = int(module_id)
        if not (await db.execute(select(Module.id).where(Module.id == module_id))).scalar_one_or_none():
            raise HTTPException(400, "Modul topilmadi")
    else:
        module_id = None

    # ── Topic (existing id OR new) ─────────────────────────────────
    topic_id = payload.get("topic_id")
    new_topic_title = (payload.get("new_topic_title") or "").strip()

    if not topic_id and not new_topic_title:
        raise HTTPException(400, "Mavzu tanlanmagan yoki kiritilmagan")

    if new_topic_title:
        topic = Topic(
            title=new_topic_title,
            module_id=module_id,
            is_active=True,
            order=0,
        )
        db.add(topic)
        await db.flush()
        topic_id = topic.id
    else:
        topic_id = int(topic_id)
        existing = (
            await db.execute(
                select(Topic).where(Topic.id == topic_id).options(lazyload(Topic.quizzes))
            )
        ).scalar_one_or_none()
        if not existing:
            raise HTTPException(400, "Mavzu topilmadi")
        # If admin chose a module and topic wasn't linked, link it
        if module_id and not existing.module_id:
            existing.module_id = module_id

    # ── Quiz ───────────────────────────────────────────────────────
    title = (payload.get("quiz_title") or "").strip()
    if not title:
        raise HTTPException(400, "Test sarlavhasi kiritilmagan")

    time_limit = int(payload.get("time_limit_seconds") or 300)

    quiz = Quiz(
        topic_id=topic_id,
        title=title,
        description=(payload.get("quiz_description") or "").strip() or None,
        time_limit_seconds=time_limit,
        is_active=True,
    )
    db.add(quiz)
    await db.flush()

    # ── Questions ──────────────────────────────────────────────────
    questions_in: list[dict] = payload.get("questions") or []
    if not questions_in:
        raise HTTPException(400, "Kamida bitta savol kerak")

    saved = 0
    for i, q in enumerate(questions_in, start=1):
        text = (q.get("text") or "").strip()
        image_url = (q.get("image_url") or "").strip()
        a = (q.get("option_a") or "").strip()
        b = (q.get("option_b") or "").strip()
        c = (q.get("option_c") or "").strip()
        d = (q.get("option_d") or "").strip()
        correct = (q.get("correct") or "A").strip().upper()

        # A question needs text OR an image, plus all four options.
        if not ((text or image_url) and a and b and c and d):
            continue
        if correct not in ("A", "B", "C", "D"):
            correct = "A"

        db.add(Question(
            quiz_id=quiz.id,
            order=i,
            text=text or None,
            image_url=image_url or None,
            option_a=a, option_b=b, option_c=c, option_d=d,
            correct_option=CorrectOption(correct),
            explanation=(q.get("explanation") or "").strip() or None,
        ))
        saved += 1

    if saved == 0:
        raise HTTPException(400, "Hech qanday to'liq savol kiritilmagan")

    # Capture ids BEFORE commit — attributes expire after commit and would
    # trigger a sync lazy-load on an async session.
    quiz_id = quiz.id
    quiz_title = quiz.title
    await db.commit()

    notify = bool(payload.get("notify", True))
    if notify:
        asyncio.create_task(notify_new_quiz(quiz_id, quiz_title, settings.WEBAPP_URL))

    return {
        "ok": True,
        "topic_id": topic_id,
        "quiz_id": quiz_id,
        "questions_saved": saved,
        "notify_scheduled": notify,
    }


# ═══ Books upload tool ═════════════════════════════════════════════

@router.get("/books", response_class=HTMLResponse)
async def books_page(request: Request, db: AsyncSession = Depends(get_db)):
    topics = (await db.execute(select(Topic).order_by(Topic.order, Topic.id))).scalars().all()
    books = (await db.execute(select(Book).order_by(Book.createdAt.desc()))).scalars().all()
    used = {b.category for b in books if b.category}
    # Fixed list first, then any extra categories admin has previously typed.
    categories = list(BOOK_CATEGORIES) + sorted(used - set(BOOK_CATEGORIES))
    return templates.TemplateResponse(
        "admin_books.html",
        {
            "request": request,
            "topics": [{"id": t.id, "title": t.title} for t in topics],
            "categories": categories,
            "default_categories": BOOK_CATEGORIES,
            "books": [
                {
                    "id": b.id,
                    "title": b.title,
                    "author": b.author,
                    "category": b.category,
                    "topic_id": b.topic_id,
                    "file_name": b.file_name,
                    "file_size": b.file_size,
                    "downloads": b.downloads,
                    "is_active": b.is_active,
                }
                for b in books
            ],
        },
    )


@router.post("/books/upload")
async def books_upload(
    db: AsyncSession = Depends(get_db),
    file: UploadFile = File(...),
    title: str = Form(...),
    author: str = Form(""),
    category: str = Form(""),
    description: str = Form(""),
    topic_id: str = Form(""),
):
    title = title.strip()
    if not title:
        raise HTTPException(400, "Sarlavha kiritilmagan")

    ext = Path(file.filename or "").suffix.lower()
    if ext not in ALLOWED_BOOK_EXT:
        raise HTTPException(400, f"Fayl turi qo'llab-quvvatlanmaydi: {ext or 'nomaʼlum'}")

    data = await file.read()
    if len(data) > MAX_BOOK_SIZE:
        raise HTTPException(400, "Fayl juda katta (maks. 100 MB)")
    if not data:
        raise HTTPException(400, "Fayl bo'sh")

    tid: int | None = None
    if topic_id.strip():
        tid = int(topic_id)
        if not await db.get(Topic, tid):
            raise HTTPException(400, "Mavzu topilmadi")

    BOOKS_DIR.mkdir(parents=True, exist_ok=True)
    stored = f"books/{uuid.uuid4().hex}{ext}"
    (MEDIA_ROOT / stored).write_bytes(data)

    book = Book(
        title=title,
        author=author.strip() or None,
        category=category.strip() or None,
        description=description.strip() or None,
        topic_id=tid,
        file_path=stored,
        file_name=file.filename or f"kitob{ext}",
        file_size=len(data),
        is_active=True,
        downloads=0,
        order=0,
    )
    db.add(book)
    await db.commit()
    return {"ok": True, "id": book.id}


@router.post("/books/{book_id}/delete")
async def books_delete(book_id: int, db: AsyncSession = Depends(get_db)):
    book = await db.get(Book, book_id)
    if not book:
        raise HTTPException(404, "Kitob topilmadi")
    try:
        (MEDIA_ROOT / book.file_path).unlink(missing_ok=True)
    except OSError:
        pass
    await db.delete(book)
    await db.commit()
    return {"ok": True}


# ═══ Question image upload ═════════════════════════════════════════

QUESTION_IMAGES_DIR = MEDIA_ROOT / QUESTION_IMAGES_DIR_NAME
SHOP_IMAGES_DIR = MEDIA_ROOT / "shop"


@router.post("/upload-image")
async def upload_image(file: UploadFile = File(...)):
    """Upload a question image and return its public URL under /media."""
    ext = Path(file.filename or "").suffix.lower()
    if ext not in ALLOWED_IMAGE_EXT:
        raise HTTPException(400, f"Rasm turi qo'llab-quvvatlanmaydi: {ext or 'nomaʼlum'}")

    data = await file.read()
    if len(data) > MAX_IMAGE_SIZE:
        raise HTTPException(400, "Rasm juda katta (maks. 5 MB)")
    if not data:
        raise HTTPException(400, "Rasm bo'sh")

    QUESTION_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    stored = f"{QUESTION_IMAGES_DIR_NAME}/{uuid.uuid4().hex}{ext}"
    (MEDIA_ROOT / stored).write_bytes(data)

    return {"ok": True, "url": f"/media/{stored}"}


@router.post("/upload-shop-image")
async def upload_shop_image(file: UploadFile = File(...)):
    """Upload a shop book cover image."""
    ext = Path(file.filename or "").suffix.lower()
    if ext not in ALLOWED_IMAGE_EXT:
        raise HTTPException(400, f"Rasm turi qo'llab-quvvatlanmaydi: {ext or 'nomaʼlum'}")

    data = await file.read()
    if len(data) > MAX_IMAGE_SIZE:
        raise HTTPException(400, "Rasm juda katta (maks. 5 MB)")
    if not data:
        raise HTTPException(400, "Rasm bo'sh")

    SHOP_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    stored = f"shop/{uuid.uuid4().hex}{ext}"
    (MEDIA_ROOT / stored).write_bytes(data)

    return {"ok": True, "url": f"/media/{stored}"}


# ─── Shop books custom management page ──────────────────────────────────

@router.get("/shop-books", response_class=HTMLResponse)
async def shop_books_page(request: Request, db: AsyncSession = Depends(get_db)):
    books = (
        await db.execute(select(ShopBook).order_by(ShopBook.order, ShopBook.id))
    ).scalars().all()
    return templates.TemplateResponse(
        "admin_shop_books.html",
        {"request": request, "books": books},
    )


@router.post("/shop-books/save")
async def shop_book_save(payload: dict[str, Any], db: AsyncSession = Depends(get_db)):
    book_id = payload.get("id")
    title = (payload.get("title") or "").strip()
    if not title:
        raise HTTPException(400, "Kitob nomi kiritilmagan")

    price = int(payload.get("price") or 0)
    description = (payload.get("description") or "").strip() or None
    cover_image_url = (payload.get("cover_image_url") or "").strip() or None
    order = int(payload.get("order") or 0)
    is_active = bool(payload.get("is_active", True))

    if book_id:
        book = await db.get(ShopBook, int(book_id))
        if not book:
            raise HTTPException(404, "Kitob topilmadi")
        book.title = title
        book.description = description
        book.price = price
        book.cover_image_url = cover_image_url
        book.order = order
        book.is_active = is_active
    else:
        book = ShopBook(
            title=title, description=description, price=price,
            cover_image_url=cover_image_url, order=order, is_active=is_active,
        )
        db.add(book)

    await db.commit()
    await db.refresh(book)
    return {"ok": True, "id": book.id}


@router.delete("/shop-books/{book_id}")
async def shop_book_delete(book_id: int, db: AsyncSession = Depends(get_db)):
    book = await db.get(ShopBook, book_id)
    if not book:
        raise HTTPException(404, "Kitob topilmadi")
    await db.delete(book)
    await db.commit()
    return {"ok": True}


# ═══ Contests (yutuqli test) ════════════════════════════════════════

def _parse_dt(value: str | None) -> datetime | None:
    if not value:
        return None
    s = value.strip().replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        raise HTTPException(400, f"Sana formati noto'g'ri: {value}")


def _contest_status(c: Contest, now: datetime) -> str:
    if not c.is_active:
        return "inactive"
    if now < c.start_at:
        return "upcoming"
    if now > c.end_at:
        return "finished"
    return "live"


@router.get("/contests", response_class=HTMLResponse)
async def contests_page(request: Request, db: AsyncSession = Depends(get_db)):
    rows = (
        await db.execute(select(Contest).order_by(desc(Contest.start_at)))
    ).scalars().all()

    counts_rows = await db.execute(
        select(ContestAttempt.contest_id, func.count(ContestAttempt.id))
        .group_by(ContestAttempt.contest_id)
    )
    counts = dict(counts_rows.all())

    now = datetime.now().astimezone()
    contests = []
    for c in rows:
        q_count = len(c.questions)
        contests.append({
            "id": c.id,
            "title": c.title,
            "prize": c.prize,
            "start_at": c.start_at.isoformat() if c.start_at else "",
            "end_at": c.end_at.isoformat() if c.end_at else "",
            "question_count": c.question_count,
            "questions_saved": q_count,
            "time_limit_seconds": c.time_limit_seconds,
            "is_active": c.is_active,
            "status": _contest_status(c, now),
            "participants": counts.get(c.id, 0),
        })
    return templates.TemplateResponse(
        "admin_contests.html",
        {"request": request, "contests": contests},
    )


@router.get("/contests/new", response_class=HTMLResponse)
async def contest_builder_page(request: Request):
    return templates.TemplateResponse(
        "admin_contest_builder.html",
        {"request": request, "contest": None},
    )


@router.get("/contests/{contest_id}/edit", response_class=HTMLResponse)
async def contest_edit_page(contest_id: int, request: Request, db: AsyncSession = Depends(get_db)):
    contest = await db.get(Contest, contest_id)
    if not contest:
        raise HTTPException(404, "Yutuqli test topilmadi")
    data = {
        "id": contest.id,
        "title": contest.title,
        "description": contest.description or "",
        "prize": contest.prize or "",
        "start_at": contest.start_at.isoformat() if contest.start_at else "",
        "end_at": contest.end_at.isoformat() if contest.end_at else "",
        "time_limit_seconds": contest.time_limit_seconds,
        "question_count": contest.question_count,
        "is_active": contest.is_active,
        "questions": [
            {
                "text": q.text or "",
                "image_url": q.image_url or "",
                "option_a": q.option_a,
                "option_b": q.option_b,
                "option_c": q.option_c,
                "option_d": q.option_d,
                "correct": q.correct_option.value,
                "explanation": q.explanation or "",
            }
            for q in sorted(contest.questions, key=lambda x: x.order)
        ],
    }
    return templates.TemplateResponse(
        "admin_contest_builder.html",
        {"request": request, "contest": data},
    )


@router.post("/contests/save")
async def contest_save(payload: dict[str, Any], db: AsyncSession = Depends(get_db)):
    contest_id = payload.get("id")

    title = (payload.get("title") or "").strip()
    if not title:
        raise HTTPException(400, "Sarlavha kiritilmagan")

    start_at = _parse_dt(payload.get("start_at"))
    end_at = _parse_dt(payload.get("end_at"))
    if not start_at or not end_at:
        raise HTTPException(400, "Boshlanish va tugash sanalari kerak")
    if end_at <= start_at:
        raise HTTPException(400, "Tugash sanasi boshlanishdan keyin bo'lishi kerak")

    question_count = int(payload.get("question_count") or 50)
    time_limit = int(payload.get("time_limit_seconds") or 3000)

    questions_in: list[dict] = payload.get("questions") or []
    valid_questions = []
    for q in questions_in:
        text = (q.get("text") or "").strip()
        image_url = (q.get("image_url") or "").strip()
        a = (q.get("option_a") or "").strip()
        b = (q.get("option_b") or "").strip()
        c_ = (q.get("option_c") or "").strip()
        d = (q.get("option_d") or "").strip()
        correct = (q.get("correct") or "A").strip().upper()
        if not ((text or image_url) and a and b and c_ and d):
            continue
        if correct not in ("A", "B", "C", "D"):
            correct = "A"
        valid_questions.append({
            "text": text or None, "image_url": image_url or None,
            "option_a": a, "option_b": b, "option_c": c_, "option_d": d,
            "correct": correct, "explanation": (q.get("explanation") or "").strip() or None,
        })

    if len(valid_questions) < question_count:
        raise HTTPException(
            400,
            f"Kamida {question_count} ta to'liq savol kerak (hozir: {len(valid_questions)})",
        )

    if contest_id:
        contest = await db.get(Contest, int(contest_id))
        if not contest:
            raise HTTPException(404, "Yutuqli test topilmadi")
        contest.title = title
        contest.description = (payload.get("description") or "").strip() or None
        contest.prize = (payload.get("prize") or "").strip() or None
        contest.start_at = start_at
        contest.end_at = end_at
        contest.time_limit_seconds = time_limit
        contest.question_count = question_count
        contest.is_active = bool(payload.get("is_active", True))
        # Replace questions
        for old in list(contest.questions):
            await db.delete(old)
        await db.flush()
        is_new = False
    else:
        contest = Contest(
            title=title,
            description=(payload.get("description") or "").strip() or None,
            prize=(payload.get("prize") or "").strip() or None,
            start_at=start_at,
            end_at=end_at,
            time_limit_seconds=time_limit,
            question_count=question_count,
            is_active=bool(payload.get("is_active", True)),
        )
        db.add(contest)
        await db.flush()
        is_new = True

    for i, q in enumerate(valid_questions, start=1):
        db.add(ContestQuestion(
            contest_id=contest.id,
            order=i,
            text=q["text"],
            image_url=q["image_url"],
            option_a=q["option_a"], option_b=q["option_b"],
            option_c=q["option_c"], option_d=q["option_d"],
            correct_option=CorrectOption(q["correct"]),
            explanation=q["explanation"],
        ))

    cid = contest.id
    ctitle = contest.title
    cprize = contest.prize
    cstart_iso = contest.start_at.strftime("%d.%m.%Y %H:%M")
    await db.commit()

    notify = bool(payload.get("notify", is_new))
    if notify and is_new:
        asyncio.create_task(
            notify_new_contest(cid, ctitle, cprize, cstart_iso, settings.WEBAPP_URL)
        )

    return {"ok": True, "id": cid, "questions_saved": len(valid_questions), "notify_scheduled": notify and is_new}


@router.post("/contests/{contest_id}/delete")
async def contest_delete(contest_id: int, db: AsyncSession = Depends(get_db)):
    contest = await db.get(Contest, contest_id)
    if not contest:
        raise HTTPException(404, "Yutuqli test topilmadi")
    await db.delete(contest)
    await db.commit()
    return {"ok": True}


@router.get("/contests/{contest_id}/winners", response_class=HTMLResponse)
async def contest_winners_page(contest_id: int, request: Request, db: AsyncSession = Depends(get_db)):
    contest = await db.get(Contest, contest_id)
    if not contest:
        raise HTTPException(404, "Yutuqli test topilmadi")

    rows = (
        await db.execute(
            select(ContestAttempt)
            .where(ContestAttempt.contest_id == contest_id)
            .options(selectinload(ContestAttempt.user))
            .order_by(desc(ContestAttempt.score), ContestAttempt.time_taken_seconds)
        )
    ).scalars().all()

    participants = []
    for i, a in enumerate(rows, start=1):
        u = a.user
        full = " ".join(filter(None, [u.first_name, u.last_name])) or u.username or f"#{u.telegram_id}"
        participants.append({
            "rank": i,
            "name": full,
            "username": u.username,
            "phone": u.phone,
            "score": a.score,
            "total": a.total,
            "percentage": a.percentage,
            "time_taken_seconds": a.time_taken_seconds,
            "completed_at": a.completed_at.strftime("%d.%m.%Y %H:%M") if a.completed_at else "",
        })

    return templates.TemplateResponse(
        "admin_contest_winners.html",
        {
            "request": request,
            "contest": {
                "id": contest.id,
                "title": contest.title,
                "prize": contest.prize,
                "start_at": contest.start_at.strftime("%d.%m.%Y %H:%M"),
                "end_at": contest.end_at.strftime("%d.%m.%Y %H:%M"),
            },
            "participants": participants,
        },
    )


@router.get("/contests/{contest_id}/export")
async def contest_export(contest_id: int, db: AsyncSession = Depends(get_db)):
    contest = await db.get(Contest, contest_id)
    if not contest:
        raise HTTPException(404, "Yutuqli test topilmadi")

    rows = (
        await db.execute(
            select(ContestAttempt)
            .where(ContestAttempt.contest_id == contest_id)
            .options(selectinload(ContestAttempt.user))
            .order_by(desc(ContestAttempt.score), ContestAttempt.time_taken_seconds)
        )
    ).scalars().all()

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Ishtirokchilar"

    headers = [
        "O'rin", "Ism Familya", "Username", "Telegram ID", "Telefon",
        "To'g'ri javob", "Jami savol", "Foiz (%)", "Vaqt (soniya)",
        "Vaqt (mm:ss)", "Tugatgan sana",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="6C63F6", end_color="6C63F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, a in enumerate(rows, start=1):
        u = a.user
        full = " ".join(filter(None, [u.first_name, u.last_name])) or ""
        mmss = f"{a.time_taken_seconds // 60:02d}:{a.time_taken_seconds % 60:02d}"
        ws.append([
            i,
            full,
            u.username or "",
            u.telegram_id,
            u.phone or "",
            a.score,
            a.total,
            a.percentage,
            a.time_taken_seconds,
            mmss,
            a.completed_at.strftime("%d.%m.%Y %H:%M") if a.completed_at else "",
        ])

    widths = [6, 28, 20, 16, 18, 14, 12, 10, 14, 12, 20]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", contest.title)[:40] or f"contest_{contest_id}"
    filename = f"{safe}_ishtirokchilar.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═══ Motivational quotes ═══════════════════════════════════════════

@router.get("/quotes", response_class=HTMLResponse)
async def quotes_page(request: Request, db: AsyncSession = Depends(get_db)):
    from models.quote import MotivationalQuote

    rows = (
        await db.execute(
            select(MotivationalQuote)
            .where(MotivationalQuote.deletedAt.is_(None))
            .order_by(desc(MotivationalQuote.id))
        )
    ).scalars().all()
    quotes = [
        {"id": q.id, "text": q.text, "is_active": q.is_active}
        for q in rows
    ]
    return templates.TemplateResponse(
        "admin_quotes.html",
        {"request": request, "quotes": quotes},
    )


@router.post("/quotes/save")
async def quotes_save(payload: dict[str, Any], db: AsyncSession = Depends(get_db)):
    from models.quote import MotivationalQuote

    qid = payload.get("id")
    text = (payload.get("text") or "").strip()
    is_active = bool(payload.get("is_active", True))
    if not text or len(text) > 2000:
        raise HTTPException(400, "Matn bo'sh yoki juda uzun (maks. 2000 belgi)")

    if qid:
        q = await db.get(MotivationalQuote, qid)
        if not q or q.deletedAt is not None:
            raise HTTPException(404, "Quote topilmadi")
        q.text = text
        q.is_active = is_active
    else:
        q = MotivationalQuote(text=text, is_active=is_active)
        db.add(q)

    await db.commit()
    await db.refresh(q)
    return {"ok": True, "id": q.id}


@router.post("/quotes/{quote_id}/delete")
async def quotes_delete(quote_id: int, db: AsyncSession = Depends(get_db)):
    from models.quote import MotivationalQuote

    q = await db.get(MotivationalQuote, quote_id)
    if not q or q.deletedAt is not None:
        raise HTTPException(404, "Quote topilmadi")
    q.deletedAt = datetime.now().astimezone()
    await db.commit()
    return {"ok": True}


# ═══ Broadcast tool ═════════════════════════════════════════════════

@router.get("/broadcast", response_class=HTMLResponse)
async def broadcast_page(request: Request, db: AsyncSession = Depends(get_db)):
    from models.telegram_user import TelegramUser
    from sqlalchemy import func

    total = await db.scalar(
        select(func.count()).select_from(TelegramUser).where(TelegramUser.is_blocked == False)  # noqa: E712
    ) or 0
    return templates.TemplateResponse(
        "admin_broadcast.html",
        {"request": request, "recipients": int(total)},
    )


@router.post("/broadcast/send")
async def broadcast_send(payload: dict[str, Any]):
    text = (payload.get("text") or "").strip()
    if not text:
        raise HTTPException(400, "Xabar matni kiritilmagan")
    if len(text) > 4000:
        raise HTTPException(400, "Xabar juda uzun (maks. 4000 belgi)")

    url = (payload.get("button_url") or "").strip() or None
    label = (payload.get("button_text") or "").strip() or None

    from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup

    kb = None
    if url and label:
        if not (url.startswith("http://") or url.startswith("https://") or url.startswith("tg://")):
            raise HTTPException(400, "Havola http(s):// yoki tg:// bilan boshlanishi kerak")
        kb = InlineKeyboardMarkup(
            inline_keyboard=[[InlineKeyboardButton(text=label, url=url)]]
        )

    stats = await broadcast(text, reply_markup=kb)
    return {"ok": True, **stats}


@router.post("/builder/parse")
async def builder_parse(payload: dict[str, Any]):
    """Parse a bulk-pasted questions block into structured questions."""
    text = payload.get("text") or ""
    return {"questions": _parse_bulk(text)}


# ── helpers ────────────────────────────────────────────────────────

_OPT_RE = re.compile(r"^\s*([A-D])[\)\.\:\-]\s*(.+?)\s*$", re.IGNORECASE)
_ANS_RE = re.compile(r"^\s*(?:answer|javob|to['`]?g['`]?ri)\s*[:\-]\s*([A-D])\s*$", re.IGNORECASE)


def _parse_bulk(raw: str) -> list[dict]:
    """
    Parse blocks like:

        1. What is 2+2?
        A) 3
        B) 4
        C) 5
        D) 6
        Answer: B

    Questions are separated by blank line or a `---` line.
    """
    blocks: list[list[str]] = []
    current: list[str] = []
    for line in raw.splitlines():
        s = line.rstrip()
        if not s.strip() or s.strip() == "---":
            if current:
                blocks.append(current)
                current = []
            continue
        current.append(s)
    if current:
        blocks.append(current)

    out: list[dict] = []
    for block in blocks:
        q: dict[str, str] = {"text": "", "option_a": "", "option_b": "",
                             "option_c": "", "option_d": "", "correct": "A"}
        text_lines: list[str] = []
        for line in block:
            m_ans = _ANS_RE.match(line)
            if m_ans:
                q["correct"] = m_ans.group(1).upper()
                continue
            m_opt = _OPT_RE.match(line)
            if m_opt:
                letter = m_opt.group(1).upper()
                q[f"option_{letter.lower()}"] = m_opt.group(2)
                continue
            text_lines.append(line)

        text = " ".join(l.strip() for l in text_lines).strip()
        # strip leading numbering like "1." / "1)"
        text = re.sub(r"^\s*\d+\s*[\.\)\-\:]\s*", "", text).strip()
        q["text"] = text

        if q["text"] and q["option_a"] and q["option_b"] and q["option_c"] and q["option_d"]:
            out.append(q)

    return out


# ═══ Shop / Orders management ══════════════════════════════════════════════

@router.get("/orders", response_class=HTMLResponse)
async def orders_page(request: Request):
    return templates.TemplateResponse("admin_orders.html", {"request": request})


@router.get("/orders/list")
async def orders_list(db: AsyncSession = Depends(get_db)):
    rows = (
        await db.execute(
            select(BookOrder)
            .options(selectinload(BookOrder.user), selectinload(BookOrder.book))
            .order_by(BookOrder.createdAt.desc())
        )
    ).scalars().all()
    return [
        {
            "id": o.id,
            "status": o.status.value,
            "user_name": o.user.full_name if o.user else "—",
            "user_phone": o.user.phone if o.user else None,
            "user_username": o.user.username if o.user else None,
            "user_telegram_id": o.user.telegram_id if o.user else None,
            "book_title": o.book.title if o.book else "—",
            "book_price": o.book.price if o.book else 0,
            "delivery_name": o.delivery_name,
            "delivery_phone": o.delivery_phone,
            "delivery_address": o.delivery_address,
            "admin_note": o.admin_note,
            "created_at": o.createdAt.isoformat() if o.createdAt else None,
        }
        for o in rows
    ]


async def _notify_order_user(telegram_id: int, order_id: int, text: str) -> None:
    """Send a bot message to the order's user and set FSM state if needed."""
    from aiogram.fsm.context import FSMContext
    from aiogram.fsm.storage.base import StorageKey
    from bot.router import BookDelivery
    from bot.setup import bot, dp

    await bot.send_message(telegram_id, text, parse_mode="HTML")

    # Put the user into the delivery-info FSM flow
    key = StorageKey(bot_id=bot.id, chat_id=telegram_id, user_id=telegram_id)
    state = FSMContext(storage=dp.storage, key=key)
    await state.set_state(BookDelivery.name)
    await state.update_data(order_id=order_id)


@router.post("/orders/{order_id}/confirm")
async def order_confirm(order_id: int, payload: dict[str, Any], db: AsyncSession = Depends(get_db)):
    order = (
        await db.execute(
            select(BookOrder)
            .where(BookOrder.id == order_id)
            .options(selectinload(BookOrder.user), selectinload(BookOrder.book))
        )
    ).scalar_one_or_none()
    if not order:
        raise HTTPException(404, "Buyurtma topilmadi")
    if order.status != OrderStatus.PENDING:
        raise HTTPException(400, f"Holat allaqachon: {order.status.value}")

    order.status = OrderStatus.CONFIRMED
    if payload.get("note"):
        order.admin_note = payload["note"]
    await db.commit()

    # Notify user via bot
    book_title = order.book.title if order.book else "kitob"
    try:
        await _notify_order_user(
            order.user.telegram_id,
            order_id,
            f"✅ <b>To'lovingiz qabul qilindi!</b>\n\n"
            f"📚 Kitob: <b>{book_title}</b>\n\n"
            "Yetkazib berish uchun quyidagi ma'lumotlarni yuboring:\n\n"
            "👤 <b>Ism Familyangizni</b> kiriting:",
        )
    except Exception:
        logger.exception("Failed to notify user %s after payment confirm", order.user.telegram_id)

    return {"ok": True}


@router.post("/orders/{order_id}/ship")
async def order_ship(order_id: int, payload: dict[str, Any], db: AsyncSession = Depends(get_db)):
    order = (
        await db.execute(
            select(BookOrder)
            .where(BookOrder.id == order_id)
            .options(selectinload(BookOrder.user), selectinload(BookOrder.book))
        )
    ).scalar_one_or_none()
    if not order:
        raise HTTPException(404, "Buyurtma topilmadi")
    if order.status != OrderStatus.PROCESSING:
        raise HTTPException(400, f"Holat noto'g'ri: {order.status.value}")

    order.status = OrderStatus.SHIPPED
    if payload.get("note"):
        order.admin_note = payload["note"]
    await db.commit()

    book_title = order.book.title if order.book else "Kitob"
    try:
        from bot.setup import bot
        await bot.send_message(
            order.user.telegram_id,
            f"📦 <b>{book_title} jo'natildi!</b>\n\n"
            "Kitobingiz tez kunda sizga yetkaziladi. 🚚\n"
            "Savollar bo'lsa adminga murojaat qiling.",
            parse_mode="HTML",
        )
    except Exception:
        logger.exception("Failed to notify user %s after ship", order.user.telegram_id)

    return {"ok": True}
